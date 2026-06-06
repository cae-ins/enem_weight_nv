#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Build POP_LFS_BY_REGION_SEX_2AGEGR_<year>_<quarter>.xlsx from ANStat TAB2.

The target format has two blocks:
  1. Regional constraints: region x sex x {0_14, 15_plus}, no Milieu.
  2. National constraints: sex x Milieu x detailed age groups.

The ANStat TAB2 sheets contain overlapping age columns:
  - "0-15"
  - "15 ans et plus" / "15+"

Therefore:
  total_all = 0-15 + 16-19 + 20-24 + ... + 65+
  0_14     = total_all - 15+
  15_19    = 15+ - (20-24 + 25-29 + ... + 65+)
  15_plus  = 15+

Region names and Domain codes are not taken from ANStat. They are copied from
the template file so the output keeps the exact names/order used by calibration
scripts.
"""

from __future__ import annotations

import argparse
import math
import re
import shutil
import sys
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import openpyxl


HEADER = ["Domain", "Région (33)", "Sexe", "Milieu", "groupe_age", "Nombre"]

TAB2_AGES = [
    "0-15",
    "16-19",
    "20-24",
    "25-29",
    "30-34",
    "35-39",
    "40-44",
    "45-49",
    "50-54",
    "55-59",
    "60-64",
    "65+",
    "15+",
]

AGES_16PLUS_DETAIL = [
    "16-19",
    "20-24",
    "25-29",
    "30-34",
    "35-39",
    "40-44",
    "45-49",
    "50-54",
    "55-59",
    "60-64",
    "65+",
]

AGES_20PLUS_DETAIL = [
    "20-24",
    "25-29",
    "30-34",
    "35-39",
    "40-44",
    "45-49",
    "50-54",
    "55-59",
    "60-64",
    "65+",
]

NATIONAL_AGE_ORDER = [
    "0_14",
    "15_19",
    "20_24",
    "25_29",
    "30_34",
    "35_39",
    "40_44",
    "45_49",
    "50_54",
    "55_59",
    "60_64",
    "65_plus",
]

NATIONAL_AGE_MAP = {
    "20_24": "20-24",
    "25_29": "25-29",
    "30_34": "30-34",
    "35_39": "35-39",
    "40_44": "40-44",
    "45_49": "45-49",
    "50_54": "50-54",
    "55_59": "55-59",
    "60_64": "60-64",
    "65_plus": "65+",
}

AGES_15PLUS_TARGET = [
    "15_19",
    "20_24",
    "25_29",
    "30_34",
    "35_39",
    "40_44",
    "45_49",
    "50_54",
    "55_59",
    "60_64",
    "65_plus",
]


@dataclass
class RegionRef:
    key: str
    domain: int
    name: str


@dataclass
class PopRow:
    domain: int | None
    region: str
    sex: int
    milieu: int | None
    age: str
    value_float: float
    value_int: int | None = None

    def as_excel_row(self) -> list[Any]:
        if self.value_int is None:
            raise ValueError("Row has not been rounded yet.")
        return [
            self.domain,
            self.region,
            self.sex,
            self.milieu,
            self.age,
            self.value_int,
        ]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build POP_LFS_BY_REGION_SEX_2AGEGR Excel files from "
            "ANStat_Trimestre TAB2 sheets."
        )
    )
    parser.add_argument("--year", type=int, required=True, help="Target/source year, e.g. 2026.")
    parser.add_argument(
        "--quarter",
        required=True,
        choices=["T1", "T2", "T3", "T4"],
        help="Target/source quarter, e.g. T1.",
    )
    parser.add_argument(
        "--anstat",
        type=Path,
        default=None,
        help=(
            "Path to ANStat_Trimestre_<year>.xlsx. Default: "
            "data/06_POPULATION_ESTIMATES/<year>/ANStat_Trimestre_<year>.xlsx"
        ),
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="ANStat TAB2 sheet name. Default: TAB2_<quarter>_<year>.",
    )
    parser.add_argument(
        "--template",
        type=Path,
        required=True,
        help=(
            "Existing POP_LFS_BY_REGION_SEX_2AGEGR xlsx used only for "
            "sheet name, formatting, exact region names, order, and Domain codes."
        ),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help=(
            "Output xlsx. Default: "
            "data/06_POPULATION_ESTIMATES/<year>/<quarter>/"
            "POP_LFS_BY_REGION_SEX_2AGEGR_<year>_<quarter>.xlsx"
        ),
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite the output file if it already exists.",
    )
    parser.add_argument(
        "--backup-existing",
        action="store_true",
        help="Rename an existing output file to *_backup_<timestamp>.xlsx before writing.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Build and validate in memory, but do not write the output xlsx.",
    )
    return parser.parse_args()


def default_anstat_path(year: int) -> Path:
    return Path("data") / "06_POPULATION_ESTIMATES" / str(year) / f"ANStat_Trimestre_{year}.xlsx"


def default_output_path(year: int, quarter: str) -> Path:
    return (
        Path("data")
        / "06_POPULATION_ESTIMATES"
        / str(year)
        / quarter
        / f"POP_LFS_BY_REGION_SEX_2AGEGR_{year}_{quarter}.xlsx"
    )


def deaccent(value: Any) -> str:
    return unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode("ascii")


def canon_region(value: Any) -> str:
    text = deaccent(value).upper()
    text = text.replace("DISTRICT AUTONOME D'ABIDJAN", "ABIDJAN")
    text = text.replace("DISTRICT AUTONOME DE YAMOUSSOUKRO", "YAMOUSSOUKRO")
    text = text.replace("ENSEMBLE", "NATIONAL")
    text = text.replace("GRANDS-PONTS", "GRAND-PONTS")
    return re.sub(r"[^A-Z0-9]", "", text)


def sex_code(value: Any) -> int | None:
    text = deaccent(value).upper()
    if text.startswith("MASC"):
        return 1
    if text.startswith("FEM"):
        return 2
    return None


def number(value: Any, context: str) -> float:
    if value is None:
        raise ValueError(f"Missing numeric value for {context}.")
    try:
        return float(value)
    except Exception as exc:
        raise ValueError(f"Invalid numeric value for {context}: {value!r}") from exc


def read_region_reference(template_path: Path) -> list[RegionRef]:
    wb = openpyxl.load_workbook(template_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    regions: list[RegionRef] = []
    seen: set[str] = set()
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1 or not any(value is not None for value in row):
            continue
        domain, region = row[0], row[1]
        if domain is None:
            continue
        key = canon_region(region)
        if key in seen:
            continue
        regions.append(RegionRef(key=key, domain=int(domain), name=str(region)))
        seen.add(key)

    wb.close()

    if len(regions) != 33:
        raise ValueError(
            f"Expected 33 regional Domain/name references in template; found {len(regions)}."
        )
    return regions


def read_tab2_source(anstat_path: Path, sheet_name: str) -> dict[tuple[str, int, int | None], dict[str, float]]:
    wb = openpyxl.load_workbook(anstat_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not found in {anstat_path}.")

    ws = wb[sheet_name]
    source: dict[tuple[str, int, int | None], dict[str, float]] = {}
    current_region: str | None = None

    for excel_row, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
        if len(row) < 43:
            continue
        if row[1] is not None:
            current_region = canon_region(row[1])
        sex = sex_code(row[3])
        if sex is None:
            continue
        if current_region is None:
            raise ValueError(f"No region label available before row {excel_row}.")

        for milieu, start_col in [(1, 4), (2, 17), (None, 30)]:
            values = {
                age: number(row[start_col + offset], f"{sheet_name} row {excel_row} {age}")
                for offset, age in enumerate(TAB2_AGES)
            }
            source[(current_region, sex, milieu)] = values

    wb.close()
    return source


def read_tab1_total_if_available(anstat_path: Path, year: int, quarter: str) -> int | None:
    sheet_name = f"TAB1_{quarter}_{year}"
    wb = openpyxl.load_workbook(anstat_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return None
    ws = wb[sheet_name]

    total_row_values: list[Any] | None = None
    current_region: str | None = None
    for row in ws.iter_rows(min_row=6, values_only=True):
        if row[1] is not None:
            current_region = str(row[1])
        if current_region == "Ensemble" and row[3] == "Total":
            total_row_values = list(row)
            break
    wb.close()

    if total_row_values is None:
        return None

    # Total block in TAB1: columns M:P (0-based 12:16).
    return int(round(sum(number(v, f"{sheet_name} Ensemble Total") for v in total_row_values[12:16])))


def validate_source_keys(
    source: dict[tuple[str, int, int | None], dict[str, float]],
    regions: Iterable[RegionRef],
) -> None:
    missing: list[tuple[str, int, int | None]] = []

    for region in regions:
        for sex in (1, 2):
            key = (region.key, sex, None)
            if key not in source:
                missing.append(key)

    for sex in (1, 2):
        for milieu in (1, 2):
            key = ("NATIONAL", sex, milieu)
            if key not in source:
                missing.append(key)

    if missing:
        raise ValueError(f"Missing TAB2 source keys: {missing}")


def total_all_from_tab2(values: dict[str, float]) -> float:
    return values["0-15"] + sum(values[age] for age in AGES_16PLUS_DETAIL)


def build_float_rows(
    source: dict[tuple[str, int, int | None], dict[str, float]],
    regions: list[RegionRef],
) -> tuple[list[PopRow], list[PopRow]]:
    regional_rows: list[PopRow] = []
    national_rows: list[PopRow] = []

    for region in regions:
        for sex in (1, 2):
            values = source[(region.key, sex, None)]
            total_all = total_all_from_tab2(values)
            regional_rows.append(
                PopRow(
                    domain=region.domain,
                    region=region.name,
                    sex=sex,
                    milieu=None,
                    age="0_14",
                    value_float=total_all - values["15+"],
                )
            )
            regional_rows.append(
                PopRow(
                    domain=region.domain,
                    region=region.name,
                    sex=sex,
                    milieu=None,
                    age="15_plus",
                    value_float=values["15+"],
                )
            )

    for sex in (1, 2):
        for milieu in (1, 2):
            values = source[("NATIONAL", sex, milieu)]
            total_all = total_all_from_tab2(values)
            derived_values = {
                "0_14": total_all - values["15+"],
                "15_19": values["15+"] - sum(values[age] for age in AGES_20PLUS_DETAIL),
            }
            for age in NATIONAL_AGE_ORDER:
                if age in derived_values:
                    value = derived_values[age]
                else:
                    value = values[NATIONAL_AGE_MAP[age]]
                national_rows.append(
                    PopRow(
                        domain=None,
                        region="NATIONAL",
                        sex=sex,
                        milieu=milieu,
                        age=age,
                        value_float=value,
                    )
                )

    return regional_rows, national_rows


def assign_largest_remainder(rows: list[PopRow], target_total: int | None = None) -> None:
    if target_total is None:
        target_total = int(round(sum(row.value_float for row in rows)))

    floors = [math.floor(row.value_float) for row in rows]
    delta = target_total - sum(floors)
    rounded = floors[:]

    if delta > 0:
        order = sorted(
            range(len(rows)),
            key=lambda i: (rows[i].value_float - floors[i], -i),
            reverse=True,
        )
        for i in order[:delta]:
            rounded[i] += 1
    elif delta < 0:
        order = sorted(
            range(len(rows)),
            key=lambda i: (rows[i].value_float - floors[i], rows[i].value_float, i),
        )
        for i in order[: abs(delta)]:
            rounded[i] -= 1

    for row, value in zip(rows, rounded):
        row.value_int = int(value)


def sum_rows(rows: Iterable[PopRow], **conditions: Any) -> int:
    total = 0
    for row in rows:
        if row.value_int is None:
            raise ValueError("Cannot sum rows before rounding.")
        keep = True
        for attr, expected in conditions.items():
            actual = getattr(row, attr)
            if isinstance(expected, (list, tuple, set)):
                keep = keep and actual in expected
            else:
                keep = keep and actual == expected
        if keep:
            total += row.value_int
    return total


def round_rows_for_consistency(regional_rows: list[PopRow], national_rows: list[PopRow]) -> None:
    # First reproduce the manual process used for T1 2026:
    # 1. Round the national detailed rows globally.
    # 2. Round the regional rows globally.
    # 3. Adjust only regional rows by a few units so regional margins match
    #    the national detail by sex x aggregated age.
    assign_largest_remainder(national_rows)
    assign_largest_remainder(regional_rows)

    for sex in (1, 2):
        for age in ("0_14", "15_plus"):
            target = (
                sum_rows(national_rows, sex=sex, age="0_14")
                if age == "0_14"
                else sum_rows(national_rows, sex=sex, age=AGES_15PLUS_TARGET)
            )
            group = [row for row in regional_rows if row.sex == sex and row.age == age]
            delta = sum_rows(group) - target

            while delta != 0:
                if delta > 0:
                    chosen = max(
                        group,
                        key=lambda row: (
                            int(row.value_int) - row.value_float,
                            int(row.value_int),
                            row.region,
                        ),
                    )
                    chosen.value_int = int(chosen.value_int) - 1
                    delta -= 1
                else:
                    chosen = max(
                        group,
                        key=lambda row: (
                            row.value_float - int(row.value_int),
                            int(row.value_int),
                            row.region,
                        ),
                    )
                    chosen.value_int = int(chosen.value_int) + 1
                    delta += 1



def build_validation_report(regional_rows: list[PopRow], national_rows: list[PopRow]) -> tuple[list[str], int]:
    lines: list[str] = []
    failures = 0

    def check(label: str, left: int, right: int) -> None:
        nonlocal failures
        delta = left - right
        status = "OK" if delta == 0 else "DELTA"
        if delta != 0:
            failures += 1
        lines.append(f"{label}: regional={left} national={right} delta={delta} {status}")

    lines.append("Core coherence checks")
    check("Total", sum_rows(regional_rows), sum_rows(national_rows))
    for sex in (1, 2):
        check(f"Sexe {sex}", sum_rows(regional_rows, sex=sex), sum_rows(national_rows, sex=sex))
    check("Age 0_14", sum_rows(regional_rows, age="0_14"), sum_rows(national_rows, age="0_14"))
    check(
        "Age 15_plus",
        sum_rows(regional_rows, age="15_plus"),
        sum_rows(national_rows, age=AGES_15PLUS_TARGET),
    )
    for sex in (1, 2):
        check(
            f"Sexe {sex} age 0_14",
            sum_rows(regional_rows, sex=sex, age="0_14"),
            sum_rows(national_rows, sex=sex, age="0_14"),
        )
        check(
            f"Sexe {sex} age 15_plus",
            sum_rows(regional_rows, sex=sex, age="15_plus"),
            sum_rows(national_rows, sex=sex, age=AGES_15PLUS_TARGET),
        )

    lines.append("")
    lines.append("National milieu checks")
    national_total = sum_rows(national_rows)
    milieu_total = sum_rows(national_rows, milieu=1) + sum_rows(national_rows, milieu=2)
    status = "OK" if milieu_total == national_total else "DELTA"
    if milieu_total != national_total:
        failures += 1
    lines.append(f"Milieu 1: {sum_rows(national_rows, milieu=1)}")
    lines.append(f"Milieu 2: {sum_rows(national_rows, milieu=2)}")
    lines.append(f"Milieu 1+2: {milieu_total} national={national_total} delta={milieu_total - national_total} {status}")

    for sex in (1, 2):
        m1 = sum_rows(national_rows, sex=sex, milieu=1)
        m2 = sum_rows(national_rows, sex=sex, milieu=2)
        sex_total = sum_rows(national_rows, sex=sex)
        status = "OK" if m1 + m2 == sex_total else "DELTA"
        if m1 + m2 != sex_total:
            failures += 1
        lines.append(
            f"Sexe {sex}: milieu1={m1} milieu2={m2} "
            f"sum={m1 + m2} sex_total={sex_total} delta={m1 + m2 - sex_total} {status}"
        )

    lines.append("")
    lines.append("National detail by age")
    for age in ["0_14"] + AGES_15PLUS_TARGET:
        lines.append(f"{age}: {sum_rows(national_rows, age=age)}")

    return lines, failures


def clear_sheet_values(ws: Any) -> None:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.value = None


def write_output(template_path: Path, output_path: Path, rows: list[PopRow]) -> None:
    wb = openpyxl.load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]

    clear_sheet_values(ws)

    for col_idx, value in enumerate(HEADER, start=1):
        ws.cell(row=1, column=col_idx).value = value

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row.as_excel_row(), start=1):
            ws.cell(row=row_idx, column=col_idx).value = value

    output_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = output_path.with_name(output_path.stem + "_tmp_write.xlsx")
    wb.save(tmp_path)
    wb.close()
    tmp_path.replace(output_path)


def prepare_output_path(output_path: Path, overwrite: bool, backup_existing: bool, dry_run: bool) -> None:
    if dry_run or not output_path.exists():
        return
    if backup_existing:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = output_path.with_name(f"{output_path.stem}_backup_{timestamp}{output_path.suffix}")
        shutil.move(str(output_path), str(backup_path))
        print(f"Backed up existing output to: {backup_path}")
        return
    if overwrite:
        return
    raise FileExistsError(
        f"Output already exists: {output_path}. Use --overwrite or --backup-existing."
    )


def main() -> int:
    args = parse_args()

    anstat_path = args.anstat or default_anstat_path(args.year)
    sheet_name = args.sheet or f"TAB2_{args.quarter}_{args.year}"
    output_path = args.output or default_output_path(args.year, args.quarter)

    if not anstat_path.exists():
        raise FileNotFoundError(f"ANStat file not found: {anstat_path}")
    if not args.template.exists():
        raise FileNotFoundError(f"Template file not found: {args.template}")

    regions = read_region_reference(args.template)
    source = read_tab2_source(anstat_path, sheet_name)
    validate_source_keys(source, regions)

    regional_rows, national_rows = build_float_rows(source, regions)
    round_rows_for_consistency(regional_rows, national_rows)

    report_lines, failures = build_validation_report(regional_rows, national_rows)
    for line in report_lines:
        print(line)

    tab1_total = read_tab1_total_if_available(anstat_path, args.year, args.quarter)
    generated_total = sum_rows(national_rows)
    if tab1_total is not None:
        delta = generated_total - tab1_total
        status = "OK" if delta == 0 else "DELTA"
        print("")
        print(f"TAB1 control total: generated={generated_total} tab1={tab1_total} delta={delta} {status}")
        if delta != 0:
            failures += 1

    if failures:
        raise RuntimeError(f"Validation failed with {failures} non-zero delta(s).")

    output_rows = regional_rows + national_rows
    if len(output_rows) != 180:
        raise RuntimeError(f"Expected 180 output rows; generated {len(output_rows)}.")

    print("")
    print(f"Rows: regional={len(regional_rows)} national={len(national_rows)} total={len(output_rows)}")
    print(f"Source: {anstat_path} / {sheet_name}")
    print(f"Template: {args.template}")
    print(f"Output: {output_path}")

    if args.dry_run:
        print("Dry run: output file not written.")
        return 0

    prepare_output_path(output_path, args.overwrite, args.backup_existing, args.dry_run)
    write_output(args.template, output_path, output_rows)
    print("Output written.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
